from json import load
import openpyxl as xl
from openpyxl.chart import BarChart,Reference

wb=xl.load_workbook("Payroll.xlsx")
sheet=wb['Payroll']
cell=sheet['a1']
# print(cell.value)

for row in range(4,sheet.max_row+1):
    # print(row)
    cell=sheet.cell(row,5)
    # print(cell.value)
    cell2=sheet.cell(row,6)
    print(cell2.value)
    answer=cell.value-cell2.value
    answer_cell=sheet.cell(row,7)
    answer_cell.value=answer


values=Reference(sheet,min_row=3,max_row=sheet.max_row,min_col=1,max_col=7)

chart=BarChart()
chart.add_data(values)
sheet.add_chart(chart,'a11')



wb.save("Payroll2.xlsx")
